import yfinance as yf  # Importing yfinance for fetching stock data
import pandas as pd  # Importing pandas for data manipulation
import numpy as np  # Importing numpy for numerical computations
import datetime  # Importing datetime for handling dates
from datetime import timedelta  # Importing timedelta for date arithmetic
import time  # Importing time for time-related operations
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

# Fetches historical stock data for the last year
def get_stock_data(symbol):
    stock = yf.Ticker(symbol)
    hist = stock.history(period="1y")
    return hist

# Calculates the Relative Strength (RS) Rating based on historical stock data
def calculate_rs_rating(hist):
    current_close = hist['Adj Close'][-1]
    close_3m = hist['Adj Close'][-63]  # Approx 3 months
    close_6m = hist['Adj Close'][-126]  # Approx 6 months
    close_9m = hist['Adj Close'][-189]  # Approx 9 months
    close_12m = hist['Adj Close'][-252]  # Approx 12 months
    rs_rating = (
        ((current_close - close_3m) / close_3m) * 40 +
        ((current_close - close_6m) / close_6m) * 20 +
        ((current_close - close_9m) / close_9m) * 20 +
        ((current_close - close_12m) / close_12m) * 20
    )
    return rs_rating

# Retrieves financial statements for the given stock symbol
def get_financials(symbol):
    stock = yf.Ticker(symbol)
    financials = stock.financials
    return financials

# Calculates earnings growth based on the net income from financial statements
def calculate_earnings_growth(financials):
    earnings = financials.loc['Net Income']
    recent_earnings = earnings[0]
    year_ago_earnings = earnings[1]
    earnings_growth = ((recent_earnings - year_ago_earnings) / year_ago_earnings) * 100
    return earnings_growth

# Calculates sales growth based on the total revenue from financial statements
def calculate_sales_growth(financials):
    sales = financials.loc['Total Revenue']
    recent_sales = sales[0]
    year_ago_sales = sales[1]
    sales_growth = ((recent_sales - year_ago_sales) / year_ago_sales) * 100
    return sales_growth

# Calculates profit margin based on the most recent net income and total revenue
def calculate_profit_margin(financials):
    sales = financials.loc['Total Revenue']
    earnings = financials.loc['Net Income']
    recent_sales = sales[0]
    recent_earnings = earnings[0]
    profit_margin = (recent_earnings / recent_sales) * 100
    return profit_margin

# Calculates return on equity (ROE) based on the most recent net income and total stockholder equity
def calculate_roe(financials):
    stock = yf.Ticker(symbol)
    balance_sheet = stock.balance_sheet
    shareholders_equity = balance_sheet.loc['Total Stockholder Equity']
    recent_earnings = financials.loc['Net Income'][0]
    recent_equity = shareholders_equity[0]
    roe = (recent_earnings / recent_equity) * 100
    return roe

# Calculates the Composite Rating for a given stock symbol
def calculate_composite_rating(symbol, rs_rating):
    financials = get_financials(symbol)
    earnings_growth = calculate_earnings_growth(financials)
    sales_growth = calculate_sales_growth(financials)
    profit_margin = calculate_profit_margin(financials)
    roe = calculate_roe(financials, symbol)

    # Normalize and weight the metrics
    composite_rating = (rs_rating * 0.25 + earnings_growth * 0.25 +
                        sales_growth * 0.25 + profit_margin * 0.125 + roe * 0.125)

    return composite_rating

def my_screener(stocklist):
    # Initialize exportList DataFrame to store stock data that meets the criteria
    exportList = pd.DataFrame(columns=["Stock", "RS_Rating", "Comp_Rating", "currentClose", "20 Day MA", "50 Day MA", "70 Day MA", "150 Day MA", "200 Day MA", "52 Week Low", "52 week High"])
    print("Initial exportList created")

    # Initialize other variables
    final = []  # List to store stocks that meet the criteria
    index = []  # List to store index values of stocks that meet the criteria
    rs = []  # List to store RS_Rating values of stocks that meet the criteria
    stocks_fit_condition = 0  # Counter for stocks that meet the criteria
    n = 0  # Counter for iteration over stocks

    # Initialize counters
    adv_w = 0  # Counter for stocks with advancing weekly close
    decl_w = 0  # Counter for stocks with declining weekly close
    adv = 0  # Counter for stocks with advancing daily close
    decl = 0  # Counter for stocks with declining daily close

    # Loop through each stock in the stock list
    for symbol in stocklist:
        print(f"Processing {symbol}")
        n += 1

        hist = get_stock_data(symbol)

        print(df)
        time.sleep(0.01)  # Pause to avoid hitting API limits

        rs_rating = calculate_rs_rating(hist)
        composite_rating = calculate_composite_rating(symbol, rs_rating)

        try:
            # Calculate advancing and declining counts based on weekly close
            currentClose = df["Adj Close"].iloc[-1] # -1 is for the last value
            lastweekClose = df["Adj Close"].iloc[-5]
            if currentClose > lastweekClose:
                adv_w += 1
            else:
                decl_w += 1
        except Exception as e:
            print(e)

        try:
            # Calculate advancing and declining counts based on daily close
            currentClose = df["Adj Close"].iloc[-1]
            yesterdayClose = df["Adj Close"].iloc[-2]
            if currentClose > yesterdayClose:
                adv += 1
            else:
                decl += 1
        except Exception as e:
            print(e)

        try:
            # Calculate turnover (Market Activity Indicator, High is Good)
            turnover = df["Volume"].iloc[-1] * df["Adj Close"].iloc[-1]
        except Exception as e:
            print(e)

        try:
            # Calculate true range for the last 10 days and last 5 days (Diff Between Prices)
            true_range_5d = (max(df["Adj Close"].iloc[-5:-1]) - min(df["Adj Close"].iloc[-5:-1]))
            true_range_10d = (max(df["Adj Close"].iloc[-10:-1]) - min(df["Adj Close"].iloc[-10:-1]))
        except Exception as e:
            print(e)

        try:
            # Calculate various moving averages
            sma = [20, 50, 70, 150, 200]
            for x in sma:
                df["SMA_" + str(x)] = round(df.iloc[:, 4].rolling(window=x).mean(), 2)

            moving_average_20 = df["SMA_20"].iloc[-1]
            moving_average_50 = df["SMA_50"].iloc[-1]
            moving_average_70 = df["SMA_70"].iloc[-1]
            moving_average_150 = df["SMA_150"].iloc[-1]
            moving_average_200 = df["SMA_200"].iloc[-1]
            low_of_52week = min(df["Adj Close"].iloc[-260:])
            high_of_52week = max(df["Adj Close"].iloc[-260:])

            try:
                moving_average_200_20 = df["SMA_200"].iloc[-32]
            except Exception:
                moving_average_200_20 = 0

            # Condition checks for stock selection
            condition_1 = True # = currentClose > moving_average_150 > moving_average_200
            condition_2 = True # = moving_average_50 > moving_average_200
            condition_3 = True # = moving_average_200 > moving_average_200_20 # 200 SMA trending up for at least 1 month (ideally 4-5 months)
            condition_4 = True # = moving_average_50 > moving_average_150 > moving_average_200
            condition_5 = True # = currentClose > moving_average_50
            condition_6 = True # = currentClose >= (1.4 * low_of_52week)      # Current Price is at least 40% above 52 week low (Many of the best are up 100-300% before coming out of consolidation)
            condition_7 = True # = currentClose >= (0.75 * high_of_52week)    # Current Price is within 25% of 52 week high
            condition_8 = True # = turnover >= 1500000                        # Turnover is larger than 1.5 million
            condition_9 = True # = true_range_10d < currentClose * 0.08       # True range in the last 10 days is less than 10% of current price
            condition_10 = True #currentClose > moving_average_20          # (optional)
            condition_11 = True #true_range_5d < currentClose * 6          # (optional) true range in the last 5 days is less than 6% of current price
            condition_12 = True #= currentClose > 10

            # Check if all conditions are met
            if all([condition_1, condition_2, condition_3, condition_4, condition_5, condition_6, condition_7, condition_8, condition_9, condition_10, condition_11, condition_12]):
                final.append(symbol)  # Add stock to final list
                index.append(n)  # Add index value to index list
                rs.append(rs_rating)  # Add RS_Rating value to rs list
                stocks_fit_condition += 1  # Increment counter for stocks that meet the criteria

                # Append stock data to exportList DataFrame
                new_row = pd.DataFrame({
                    'Stock': [symbol], 
                    "RS_Rating": [rs_rating],
                    "Comp_Rating": [composite_rating],
                    "currentClose": [currentClose],
                    "20 Day MA": [moving_average_20],
                    "50 Day MA": [moving_average_50],
                    "70 Day MA": [moving_average_70],
                    "150 Day MA": [moving_average_150],
                    "200 Day MA": [moving_average_200],
                    "52 Week Low": [low_of_52week],
                    "52 week High": [high_of_52week]
                })
                exportList = pd.concat([exportList, new_row], ignore_index=True)
                print(symbol + " made the requirements")
                print(date_study)

        except Exception as e:
            print(e)
            print("No data on " + symbol)

    # Sort the exportList by RS_Rating in descending order
    exportList = exportList.sort_values(by="RS_Rating", ascending=False)
    print("Sorted exportList by RS_Rating:", exportList)

    # Save the results to CSV
    try:
        exportList.to_csv('elseOutput/stocks.csv', index=False)
        print("CSV file created successfully")
    except Exception as e:
        print(e)
        print("Failed to create CSV file")

    return exportList

def detect_vcp_pattern(stock, rs_rating, composite_rating, window=5, volume_increase=2, price_increase=1.1):
    data = get_stock_data(symbol)

    # Ensure Date is a column in the DataFrame
    data.reset_index(inplace=True)
    data["Comp_Rating"] = composite_rating
    data["Comp_Rating>50"] = composite_rating > 50

    data["RS_Rating"] = rs_rating
    data["RS_Rating>55"] = rs_rating > 55

    # Calculate moving averages for price and volume
    data["MA_Price"] = data["Adj Close"].rolling(window=window).mean()
    data["MA_Volume"] = data["Volume"].rolling(window=window).mean()
    
    # Detect periods of relatively low volatility (price stable around MA_Price)
    data["Volatility"] = np.abs(data["Adj Close"] - data["MA_Price"]) / data["MA_Price"]
    data["Volatility<0.075"] = np.abs(data["Adj Close"] - data["MA_Price"]) / data["MA_Price"] < 0.075
    
    # Detect significant volume increase
    data["Volume_Increase"] = data["Volume"] / data["MA_Volume"]
    data["Volume_Increase>" + str(volume_increase)] = data["Volume_Increase"] > volume_increase
    
    # Detect significant price increase
    data["Price_Increase"] = data["Adj Close"] / data["Adj Close"].shift(1)
    data["Price_Increase>" + str(price_increase)] = data["Price_Increase"] > price_increase
    
    # Combine conditions to detect pre-burst pattern
    data["Pre_Burst_Pattern"] = data["Volatility<0.075"]
    data["Breakout"] = data["Price_Increase>" + str(price_increase)] & data["Volume_Increase>" + str(volume_increase)]

    return data

def main():
    # Define the stock list and date for the study
    #stocklist = ["AAPL", "MSFT", "GOOGL", "TMDX", "HIMS", "SNX", "STEP", "AMG", "ANET", "ASR", "EURN", "GBDC", "HASI", "MAIN", "NVO", "OLED", "SPG", "UE", "UTHR", "VRTX", "WPM"]
    stocklist = ["HIMS"]
    exportList = my_screener(stocklist)
    
    for index, row in exportList.iterrows():
        stock = row["Stock"]
        data = detect_vcp_pattern(stock, row["RS_Rating"], row["Comp_Rating"])
        print(stock, ":", data)

         # Get dates where the pattern was detected
        pattern_dates = data[data["Breakout"]]["Date"]

        #plotData(data, pattern_dates)

        # Save to Excel with the first row and first column frozen, adjusted column widths, and defined as a table
        output_filename = f"elseOutput/{stock}.xlsx"
        with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
            data.to_excel(writer, index=False, sheet_name='PatternDates')
            worksheet = writer.sheets['PatternDates']
            worksheet.freeze_panes = 'B2'

            # Adjust column widths
            for col in worksheet.columns:
                max_length = 0
                column = col[0].column_letter  # Get the column name
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = max_length + 2  # Adding a little extra space
                worksheet.column_dimensions[column].width = adjusted_width

            # Define the table
            tab = Table(displayName="PatternTable", ref=f"A1:{get_column_letter(worksheet.max_column)}{worksheet.max_row}")

            # Add a default style with striped rows and banded columns
            style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
            tab.tableStyleInfo = style
            worksheet.add_table(tab)

        print(f"Pattern dates saved to {output_filename}")

main()
