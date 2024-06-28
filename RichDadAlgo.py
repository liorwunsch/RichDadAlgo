import sys
import pickle
import yfinance as yf
import pandas as pd 
import numpy as np
import math
import datetime
from datetime import timedelta
import pytz
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import os
import shutil

def savePickle(pickle_file_path, data):
    pickle_file_path = 'Params/' + pickle_file_path
    with open(pickle_file_path, 'wb') as file:
        pickle.dump(data, file)
        print("Saving " + pickle_file_path)

def loadPickle(pickle_file_path):
    pickle_file_path = 'Params/' + pickle_file_path
    if os.path.exists(pickle_file_path):
       with open(pickle_file_path, 'rb') as file:
           data = pickle.load(file)
           print("Loading " + pickle_file_path)
           return True, data
    else:
        return False, []

def readStockList():
    file_name = 'stocks.txt'
    with open(file_name, 'r') as file:
        stocklist = file.readlines()
    stocklist = [line.strip() for line in stocklist] # Remove newline characters from each line
    stocklist = list(set(stocklist))
    stocklist.sort()
    return stocklist

def createOutputFolder():
    folder_name = "Output"
    try:
        if os.path.exists(folder_name):
            shutil.rmtree(folder_name)
        os.makedirs(folder_name)
    except:
        print("CANNOT OPEN FOLDER: ", folder_name)
        return False
    return True

def createParamsFolder():
    folder_name = "Params"
    try:
        os.makedirs(folder_name)
    except:
        pass
    return True

def printToExcel(file_name, results=pd.DataFrame()):
    if results.empty:
        print("printToExcel: " + file_name + "'s results is empty")
        return

    output_filename = f"Output/{file_name}.xlsx"
    with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:  
        results.to_excel(writer, index=False, sheet_name='results')
        worksheet = writer.sheets['results']
        worksheet.freeze_panes = 'B2'

        # Adjust column widths
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = max_length + 2  # Adding a little extra space
            worksheet.column_dimensions[column].width = adjusted_width

        # Define the table
        tab = Table(displayName="results", ref=f"A1:{get_column_letter(worksheet.max_column)}{worksheet.max_row}")

        # Add a default style with striped rows and banded columns
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tab.tableStyleInfo = style
        worksheet.add_table(tab)
    print(f"results saved to {output_filename}")

def getLastCloseDate(tz, ny_close_hour=16, ny_close_minute=0):
    tz1 = pytz.timezone(tz)
    tz2 = pytz.timezone('America/New_York')
    _today = datetime.datetime.now()
    time1 = tz1.localize(_today)
    time2 = tz2.localize(_today)
    time_difference = (time2 - time1).total_seconds() / 3600
    time_difference_hours = math.floor(time_difference)
    time_difference_minutes = (int)((time_difference - time_difference_hours) * 60)

    # NYSE and NASDAQ After Hours at TimeZone tz
    close_time = datetime.datetime(_today.year, _today.month, _today.day, ny_close_hour, ny_close_minute)
    local_close_time = close_time + timedelta(hours=time_difference_hours, minutes=(time_difference_minutes + 1))

    if _today < local_close_time:
        _today = local_close_time - timedelta(days=1)
    else:
        _today = local_close_time

    _today = _today.replace(hour=0, minute=0)
    #print(f"last_close_date = {_today}")
    return _today

def determineActiveDates(symbol, start_date, end_date):
    data = yf.Ticker(symbol)
    hist = data.history(start=start_date, end=end_date+timedelta(days=1)) # +1 for data of today's afterhours and not yesterday's
    hist = hist.drop_duplicates()

    active_dates = hist.index
    active_dates = active_dates.tolist()
    #print(f"end_date = {end_date}")
    #print(f"active_dates = {active_dates}")
    #sys.exit()
    return active_dates

def calcCurrentMovingAverages(hist):
    sma = [20, 50, 70, 150, 200]
    for x in sma:
        hist["SMA_" + str(x)] = hist["Close"].rolling(window=x).mean()
    
    ma_20 = hist["SMA_20"].iloc[-1]
    ma_50 = hist["SMA_50"].iloc[-1]
    ma_70 = hist["SMA_70"].iloc[-1]
    ma_150 = hist["SMA_150"].iloc[-1]
    ma_200 = hist["SMA_200"].iloc[-1]
    ma_200_20 = hist["SMA_200"].iloc[-22]

    return ma_20, ma_50, ma_70, ma_150, ma_200, ma_200_20

def isMinerviniTrendTemplate(hist):
    close = hist["Close"].iloc[-1] # -1 is for the last value

    ma_20, ma_50, ma_70, ma_150, ma_200, ma_200_20 = calcCurrentMovingAverages(hist)

    high_of_52week = max(hist["Close"].iloc[-252:])
    low_of_52week = min(hist["Close"].iloc[-252:])

    # Condition checks for stock selection
    condition_1_1 = close > ma_150                 # Price > 150-day (30-week) MA
    condition_1_2 = close > ma_200                 # Price > 200-day (40-week) MA
    condition_2 = ma_150 > ma_200                  # 150-day MA > 200-day MA
    condition_3 = ma_200 > ma_200_20               # 200-day MA trending up for at least 1 month (22 days)
    condition_4_1 = ma_50 > ma_150                 # 50-day MA > 150-day MA
    condition_4_2 = ma_50 > ma_200                 # 50-day MA > 200-day MA
    condition_5 = close > ma_50                    # Price above 50-day MA
    condition_6 = close >= (0.75 * high_of_52week) # Price within 25% of 52-week high
    condition_7 = close >= (1.25 * low_of_52week)  # Price at least 25% above 52-week low

    is_trend_template = condition_1_1 & condition_1_2 & condition_2 & condition_3 & condition_4_1 & condition_4_2 & condition_5 & condition_6 & condition_7
    return is_trend_template

def isVCP(hist, _window=5, volume_increase_condition=1.75, price_increase_condition=1.05):
    # Ensure Date is a column in the DataFrame
    hist.reset_index(inplace=True)

    # Calculate moving averages for price and volume
    hist["MA_Price"] = hist["Close"].rolling(window=_window).mean()
    hist["MA_Volume"] = hist["Volume"].rolling(window=_window).mean()
    
    # Detect periods of relatively low volatility (price stable around MA_Price)
    hist["Volatility"] = np.abs(hist["Close"] - hist["MA_Price"]) / hist["MA_Price"]
    hist["Volatility<0.075"] = hist["Volatility"] < 0.075
    
    # Detect significant volume increase
    hist["Volume_Increase"] = hist["Volume"] / hist["MA_Volume"]
    hist["Volume_Increase>" + str(volume_increase_condition)] = hist["Volume_Increase"] > volume_increase_condition
    
    # Detect significant price increase
    hist["Price_Increase"] = hist["Close"] / hist["Close"].shift(1)
    hist["Price_Increase>" + str(price_increase_condition)] = hist["Price_Increase"] > price_increase_condition
    
    # Combine conditions to detect pre-burst (vcp) pattern
    hist["Pre_Burst_Pattern"] = hist["Volatility<0.075"]
    hist["Breakout"] = hist["Price_Increase>" + str(price_increase_condition)] & hist["Volume_Increase>" + str(volume_increase_condition)]

    is_vcp_pattern = hist["Pre_Burst_Pattern"].iloc[-1]
    is_vcp_breakout = hist["Breakout"].iloc[-1]
    return is_vcp_pattern

def getKpiAtDay(hist, date_study):
    index = hist.index[hist['Date'] <= date_study].tolist()
    index = index[-1]
    hist = hist.iloc[:index + 1] # +1 for data of today's afterhours and not yesterday's
    hist = hist.copy()

    is_trend_template = isMinerviniTrendTemplate(hist)
    is_vcp_pattern = isVCP(hist)

    hist_weekly = hist.resample('W', on='Date').agg({
        'Open': 'first',
        'High': 'max',
        'Low': 'min',
        'Close': 'last',
        'Volume': 'sum'
    }).dropna().reset_index()
    is_vcp_pattern_weekly = isVCP(hist_weekly)

    hist_monthly = hist.resample('ME', on='Date').agg({
        'Open': 'first',
        'High': 'max',
        'Low': 'min',
        'Close': 'last',
        'Volume': 'sum'
    }).dropna().reset_index()
    is_vcp_pattern_monthly = isVCP(hist_monthly)

    kpi_row = {
        "Date": date_study,
        "Close": hist["Close"].iloc[-1],
        "High": hist["High"].iloc[-1],
        "isTrendTemplate": is_trend_template,
        "isVcpPattern": is_vcp_pattern,
        "isVcpPattern_Weekly": is_vcp_pattern_weekly,
        "isVcpPattern_Monthly": is_vcp_pattern_monthly
    }
    return kpi_row

def getKpiAtPeriod(symbol, start_date, end_date):
    kpi_results_path = symbol + '_kpi_results.pkl'
    filtered_kpi_results_path = symbol + '_kpi_results_filtered.pkl'
    grouped_kpi_results_path = symbol + '_kpi_results_grouped.pl'
    success_1, kpi_results = loadPickle(kpi_results_path)
    success_2, filtered_kpi_results = loadPickle(filtered_kpi_results_path)
    success_3, grouped_kpi_results = loadPickle(grouped_kpi_results_path)
    if success_1 and success_2 and success_3:
        printToExcel(symbol + "_grouped", grouped_kpi_results)
        printToExcel(symbol + "_filtered", filtered_kpi_results)
        return kpi_results, filtered_kpi_results

    data = yf.Ticker(symbol)
    hist = data.history(period="max")
    hist = hist.drop_duplicates()
    hist.reset_index(inplace=True) # Ensure Date is a column in the DataFrame

    kpi_rows = []
    active_dates = determineActiveDates(symbol, start_date, end_date)

    if len(active_dates) == 0 or (hist['Date'].iloc[0] > active_dates[0] - timedelta(days=365)):
        kpi_results = pd.DataFrame()
        grouped_kpi_results = pd.DataFrame()
        filtered_kpi_results = pd.DataFrame()
    else:
        for date_study in active_dates:
            print(f"Processing {symbol} at {date_study} EST")

            kpi_row = getKpiAtDay(hist, date_study)
            kpi_rows.append(kpi_row)

        kpi_results = pd.DataFrame(kpi_rows)
        if not kpi_results.empty:
            kpi_results['Date'] = kpi_results['Date'].dt.tz_localize(None)

        prev_row = None
        grouped_kpi_rows = []
        for index, row in kpi_results.iterrows():
            if index == 0:
                index_begin = index
                date_begin = row['Date']
            else:
                if row['isTrendTemplate'] == prev_row['isTrendTemplate'] and \
                    row['isVcpPattern'] == prev_row['isVcpPattern'] and \
                    row['isVcpPattern_Weekly'] == prev_row['isVcpPattern_Weekly'] and \
                    row['isVcpPattern_Monthly'] == prev_row['isVcpPattern_Monthly']:
                    if index == kpi_results.index[-1]:
                        index_end = index
                        date_end = row['Date']
                        period_high_index = kpi_results['High'].iloc[index_begin:index_end+1].idxmax()
                        period_high_date = kpi_results['Date'].iloc[period_high_index]
                        period_high_price = kpi_results['High'].iloc[period_high_index]
                        grouped_kpi_row = {
                            "Date_Begin": date_begin,
                            "Date_End": date_end,
                            "PeriodHigh_Date": period_high_date,
                            "PeriodHigh_Price": period_high_price,
                            "isTrendTemplate": row['isTrendTemplate'],
                            "isVcpPattern": row['isVcpPattern'],
                            "isVcpPattern_Weekly": row['isVcpPattern_Weekly'],
                            "isVcpPattern_Monthly": row['isVcpPattern_Monthly']
                        }
                        grouped_kpi_rows.append(grouped_kpi_row)
                else:
                    index_end = index - 1
                    date_end = prev_row['Date']
                    period_high_index = kpi_results['High'].iloc[index_begin:index_end+1].idxmax()
                    period_high_date = kpi_results['Date'].iloc[period_high_index]
                    period_high_price = kpi_results['High'].iloc[period_high_index]
                    grouped_kpi_row = {
                        "Date_Begin": date_begin,
                        "Date_End": date_end,
                        "PeriodHigh_Date": period_high_date,
                        "PeriodHigh_Price": period_high_price,
                        "isTrendTemplate": prev_row['isTrendTemplate'],
                        "isVcpPattern": prev_row['isVcpPattern'],
                        "isVcpPattern_Weekly": prev_row['isVcpPattern_Weekly'],
                        "isVcpPattern_Monthly": prev_row['isVcpPattern_Monthly']
                    }
                    grouped_kpi_rows.append(grouped_kpi_row)
                    index_begin = index
                    date_begin = row['Date']
                    if index == kpi_results.index[-1]:
                        index_end = index
                        date_end = row['Date']
                        period_high_index = kpi_results['High'].iloc[index_begin:index_end+1].idxmax()
                        period_high_date = kpi_results['Date'].iloc[period_high_index]
                        period_high_price = kpi_results['High'].iloc[period_high_index]
                        grouped_kpi_row = {
                            "Date_Begin": date_begin,
                            "Date_End": date_end,
                            "PeriodHigh_Date": period_high_date,
                            "PeriodHigh_Price": period_high_price,
                            "isTrendTemplate": row['isTrendTemplate'],
                            "isVcpPattern": row['isVcpPattern'],
                            "isVcpPattern_Weekly": row['isVcpPattern_Weekly'],
                            "isVcpPattern_Monthly": row['isVcpPattern_Monthly']
                        }
                        grouped_kpi_rows.append(grouped_kpi_row)
            prev_row = row

        grouped_kpi_results = pd.DataFrame(grouped_kpi_rows)
        indexes = []
        if not grouped_kpi_results.empty:
            indexes = grouped_kpi_results.index[grouped_kpi_results['isTrendTemplate'] & \
                                            grouped_kpi_results['isVcpPattern'] & \
                                            grouped_kpi_results['isVcpPattern_Weekly'] & \
                                            grouped_kpi_results['isVcpPattern_Monthly']].tolist()
        filtered_kpi_results = grouped_kpi_results.iloc[indexes]

    savePickle(kpi_results_path, kpi_results)
    savePickle(filtered_kpi_results_path, filtered_kpi_results)
    savePickle(grouped_kpi_results_path, grouped_kpi_results)
    printToExcel(symbol + "_grouped", grouped_kpi_results)
    printToExcel(symbol + "_filtered", filtered_kpi_results)
    return kpi_results, filtered_kpi_results

def determineBuyPoints(kpi_results, filtered_kpi_results):
    kpi_results.loc[:, 'isBuyPoint'] = False
    for index, row in filtered_kpi_results.iterrows():
        date_end = row['Date_End']
        buy_price_date = row['PeriodHigh_Date']
        buy_price_limit = row['PeriodHigh_Price']
        date_indexes = kpi_results.index[kpi_results['Date'] > buy_price_date].tolist()
        if len(date_indexes) != 0:
            date_index = date_indexes[0]
            limit_indexes = kpi_results.index[kpi_results['High'] > buy_price_limit].tolist()
            limit_indexes = [x for x in limit_indexes if x >= date_index]
            if len(limit_indexes) != 0:
                limit_index = limit_indexes[0]
                kpi_results.loc[limit_index, 'isBuyPoint'] = True
                kpi_results.loc[limit_index, 'LimitPriceDate'] = buy_price_date
                kpi_results.loc[limit_index, 'LimitBuyingPrice'] = buy_price_limit
    return kpi_results

def addDatePriceEntries(kpi_results, date_price_entries, ispoint_col, price_col):
    # Initialize columns if not already present
    if ispoint_col not in kpi_results.columns:
        kpi_results[ispoint_col] = False
    if price_col not in kpi_results.columns:
        kpi_results[price_col] = None

    # Update columns
    for date, buy_price in date_price_entries:
        kpi_results.loc[kpi_results["Date"] == date, ispoint_col] = True
        kpi_results.loc[kpi_results["Date"] == date, price_col] = buy_price

def determineSellPoints(kpi_results, r): # r(isk)
    actual_buy_points = []
    whole_sell_points = []
    partial_sell_points_1 = []
    partial_sell_points_2 = []

    is_bought, is_partially_sold = False, False
    for index, row in kpi_results.iterrows():
        date, close = row['Date'], row['Close']
        if not is_bought:
            if row['isBuyPoint']:
                buy_price = row['LimitBuyingPrice']
                actual_buy_points.append((date, buy_price))

                stop_price = buy_price * (1 - r) # Loss (-r)
                limit_price = buy_price * (1 + 3*r) # Profit (3r)

                is_bought = True
        else:
            # sell below
            if close <= stop_price:
                if not is_partially_sold:
                    whole_sell_points.append((date, stop_price * 1)) # Loss (-r)
                else:
                    partial_sell_points_2.append((date, stop_price * 1/3)) # 2nd 'Loss' (1.5r * 1/3 = 0.5r)

                is_bought, is_partially_sold = False, False

            # sell above
            if close >= limit_price:
                if not is_partially_sold:
                    partial_sell_points_1.append((date, limit_price * 2/3)) # Profit (3r * 2/3 = 2r)
                    # buy_price stays the same
                    stop_price = buy_price * (1 + 1.5*r) # 2nd 'Loss' (1.5r * 1/3 = 0.5r)
                    limit_price = buy_price * (1 + 4.5*r) # 2nd Profit (4.5r * 1/3 = 1.5r)

                    is_partially_sold = True
                else:
                    partial_sell_points_2.append((date, limit_price * 1/3)) # 2nd Profit (4.5r * 1/3 = 1.5r)

                    is_bought, is_partially_sold = False, False

    addDatePriceEntries(kpi_results, actual_buy_points, "isBuyPointActual", "BuyPrice")
    addDatePriceEntries(kpi_results, whole_sell_points, "isSellPointWhole", "SellPrice1.00")
    addDatePriceEntries(kpi_results, partial_sell_points_1, "isSellPointPartial1", "SellPrice0.67")
    addDatePriceEntries(kpi_results, partial_sell_points_2, "isSellPointPartial2", "SellPrice0.33")
    return kpi_results

def addStockTransactions(symbol, kpi_results, r): # r(isk)
    transactions = []
    viable_entries = kpi_results[(kpi_results["isBuyPointActual"] | kpi_results["isSellPointWhole"] |
                                   kpi_results["isSellPointPartial1"] | kpi_results["isSellPointPartial2"])]
    
    for index in range(len(viable_entries)):
        row = viable_entries.iloc[index]
        if row["isBuyPointActual"]:
            buy_price = row["BuyPrice"]
            limit_price_date = row["LimitPriceDate"]
            buy_date = row["Date"]

            sell_price = -1

            try:
                next_row = viable_entries.iloc[index + 1]
                if next_row["isSellPointWhole"]:
                    sell_price = next_row["SellPrice1.00"]
                    sell_date = next_row["Date"]
                else:
                    second_next_row = viable_entries.iloc[index + 2]
                    if next_row["isSellPointPartial1"] and second_next_row["isSellPointPartial2"]:
                        sell_price = next_row["SellPrice0.67"] + second_next_row["SellPrice0.33"]
                        sell_date = second_next_row["Date"]
            except:
                print("index=", index-1, ", len=", len(viable_entries))

            if sell_price != -1:
                profit_percent = (sell_price / buy_price) - 1
                profit_r = profit_percent / r
                tax_percent = profit_percent * (25/100)

                transaction_row = {
                    "Symbol": symbol,
                    "LimitPriceDate": limit_price_date,
                    "Buy_Date": buy_date,
                    "Sell_Date": sell_date,
                    "Profit[%]": profit_percent * 100,
                    "Profit_r": profit_r,
                    "Tax[%]": tax_percent * 100
                }
                transactions.append(transaction_row)

    transactions = pd.DataFrame(transactions)
    if not transactions.empty:
        transactions['Buy_Date'] = transactions['Buy_Date'].dt.tz_localize(None)
        transactions['Sell_Date'] = transactions['Sell_Date'].dt.tz_localize(None)
    return transactions

def main(stocklist, r, num_days_back, tz):
    if not stocklist:
        stocklist = readStockList()
    if not createOutputFolder():
        return
    if not createParamsFolder():
        return

    #_today = getLastCloseDate(tz)
    _today = datetime.datetime(year=2024, month=6, day=25) # TODO
    start_date = _today - timedelta(days=num_days_back+1) # +1 for isVcpPattern of oldest entry
    end_date = _today

    #start_date = datetime.datetime(year=2024, month=2, day=1)

    transactions = pd.DataFrame()
    for symbol in stocklist:
        kpi_results, filtered_kpi_results = getKpiAtPeriod(symbol, start_date, end_date)
        printToExcel(symbol + "_kpi", kpi_results)
        if not filtered_kpi_results.empty:
            kpi_results = determineBuyPoints(kpi_results, filtered_kpi_results)
            kpi_results = determineSellPoints(kpi_results, r)
            printToExcel(symbol + "_kpi", kpi_results)
            symbol_transactions = addStockTransactions(symbol, kpi_results, r)
            printToExcel(symbol + "_transactions", symbol_transactions)
            if transactions.empty:
                transactions = symbol_transactions
            else:
                transactions = pd.concat([transactions, symbol_transactions], ignore_index=True)

    transactions_summary = pd.DataFrame({
        "Symbol": 'Summary',
        "LimitPriceDate": pd.NaT,
        "Buy_Date": pd.NaT,
        "Sell_Date": pd.NaT,
        "Profit[%]": transactions["Profit[%]"].sum(),
        "Profit_r": transactions["Profit_r"].sum(),
        "Tax[%]": transactions["Tax[%]"].sum()
    }, index=[0])
    transactions = pd.concat([transactions, transactions_summary], ignore_index=True)
    printToExcel("zzTransactions", transactions)

stocklist = []
#stocklist = ["AAPL","HEI","HIMS","WMT","ELAN","EMR","TRI","MKL","QTWO","TMDX","PINS","TGT","USFD","AMZN","META"]
r = 5/100
num_days_back = 365
tz = "Asia/Jerusalem"

main(stocklist, r, num_days_back, tz)
